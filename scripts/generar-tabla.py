#!/usr/bin/env python3
"""Convierte un CSV (p. ej. el de tarifas) en un Excel con formato (openpyxl).

- Encabezados en negrita con fondo.
- Anchos de columna automáticos.
- Si existen columnas 'tarifa' + 'hotel' + 'check_in', resalta en verde la tarifa más
  baja por (hotel, check_in) — útil para rate shopping / paridad.

Dependencias: openpyxl.

Uso:
    python scripts/generar-tabla.py reports/2026-06-07-tarifas.csv
    python scripts/generar-tabla.py entrada.csv --salida reports/tabla.xlsx --titulo "Tarifas"
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    sys.exit("Falta 'openpyxl'. Instala con: pip install -r scripts/requirements.txt")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
MIN_FILL = PatternFill("solid", fgColor="C6EFCE")  # verde claro: tarifa más baja


def _to_float(valor: str) -> float | None:
    s = (valor or "").strip().replace(".", "").replace(",", ".")
    s = "".join(ch for ch in s if ch.isdigit() or ch == ".")
    try:
        return float(s) if s else None
    except ValueError:
        return None


def leer_csv(ruta: Path) -> tuple[list[str], list[dict]]:
    if not ruta.is_file():
        sys.exit(f"No se encontró el CSV: {ruta}")
    with ruta.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        filas = list(reader)
        encabezados = reader.fieldnames or []
    if not filas:
        sys.exit(f"El CSV está vacío: {ruta}")
    return encabezados, filas


def resaltar_minimos(filas: list[dict], encabezados: list[str]) -> set[int]:
    """Devuelve los índices de fila (base 0 en datos) con la tarifa mínima por grupo."""
    if not {"tarifa", "hotel", "check_in"}.issubset(encabezados):
        return set()
    mejor: dict[tuple[str, str], tuple[float, int]] = {}
    for i, fila in enumerate(filas):
        val = _to_float(fila.get("tarifa", ""))
        if val is None:
            continue
        clave = (fila.get("hotel", ""), fila.get("check_in", ""))
        if clave not in mejor or val < mejor[clave][0]:
            mejor[clave] = (val, i)
    return {idx for _, idx in mejor.values()}


def generar(ruta_csv: Path, ruta_xlsx: Path, titulo: str) -> None:
    encabezados, filas = leer_csv(ruta_csv)
    minimos = resaltar_minimos(filas, encabezados)
    col_tarifa = encabezados.index("tarifa") + 1 if "tarifa" in encabezados else None

    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    for c, nombre in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=c, value=nombre)
        celda.fill = HEADER_FILL
        celda.font = HEADER_FONT
        celda.alignment = Alignment(horizontal="center")

    for r, fila in enumerate(filas, start=2):
        for c, nombre in enumerate(encabezados, 1):
            ws.cell(row=r, column=c, value=fila.get(nombre, ""))
        if col_tarifa and (r - 2) in minimos:
            ws.cell(row=r, column=col_tarifa).fill = MIN_FILL

    # Anchos automáticos (acotados).
    for c, nombre in enumerate(encabezados, 1):
        ancho = max(len(nombre), *(len(str(fila.get(nombre, ""))) for fila in filas))
        ws.column_dimensions[get_column_letter(c)].width = min(max(ancho + 2, 10), 60)

    ws.freeze_panes = "A2"
    ruta_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta_xlsx)
    print(f"Excel generado: {ruta_xlsx} ({len(filas)} filas)")
    if minimos:
        print(f"  Resaltadas {len(minimos)} tarifas mínimas por hotel/fecha (verde).")


def main() -> None:
    parser = argparse.ArgumentParser(description="CSV -> Excel con formato.")
    parser.add_argument("csv", help="Ruta del CSV de entrada.")
    parser.add_argument("--salida", help="Ruta del .xlsx (por defecto, junto al CSV).")
    parser.add_argument("--titulo", default="Datos", help="Título de la hoja.")
    args = parser.parse_args()

    ruta_csv = Path(args.csv)
    ruta_xlsx = Path(args.salida) if args.salida else ruta_csv.with_suffix(".xlsx")
    generar(ruta_csv, ruta_xlsx, args.titulo)


if __name__ == "__main__":
    main()
