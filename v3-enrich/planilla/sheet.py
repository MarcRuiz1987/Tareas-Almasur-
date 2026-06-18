"""Lectura/escritura de planillas (.xlsx / .csv) con mapeo tolerante de columnas.

La planilla se carga a memoria como una lista de filas; se detecta la fila de
encabezados y se mapean los **campos lógicos** (ver ``config.ALIAS_COLUMNAS``) a
índices de columna, tolerando mayúsculas, acentos y nombres alternativos. Así la
v3 funciona con planillas heterogéneas sin tocar el código.

Reglas de oro al completar:
  - sólo se rellenan celdas **vacías** (salvo que se pida ``sobrescribir``),
  - nunca se borran columnas ni filas existentes,
  - las columnas que faltan se **agregan** al final con su encabezado canónico.
"""

from __future__ import annotations

import csv
import unicodedata
from pathlib import Path

from .config import ALIAS_COLUMNAS, ENCABEZADOS_CANONICOS


def _normalizar(texto: object) -> str:
    """minúsculas, sin acentos y sin espacios sobrantes — para comparar encabezados."""
    s = str(texto or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.split())


def _vacio(valor: object) -> bool:
    return valor is None or str(valor).strip() == ""


class Planilla:
    """Una planilla cargada en memoria, con mapeo campo_lógico → columna."""

    def __init__(self, encabezados: list[str], filas: list[list], origen: Path):
        self.encabezados = encabezados
        self.filas = filas  # cada fila es una lista alineada con encabezados
        self.origen = origen
        self.mapa = self._mapear_columnas()

    # ─── Carga / guardado ────────────────────────────────────────────────────

    @classmethod
    def cargar(cls, path: str | Path) -> "Planilla":
        path = Path(path)
        if path.suffix.lower() in (".xlsx", ".xlsm"):
            encabezados, filas = cls._leer_xlsx(path)
        elif path.suffix.lower() == ".csv":
            encabezados, filas = cls._leer_csv(path)
        else:
            raise ValueError(f"Formato no soportado: {path.suffix} (usa .xlsx o .csv)")
        return cls(encabezados, filas, path)

    @staticmethod
    def _leer_xlsx(path: Path) -> tuple[list[str], list[list]]:
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True)
        ws = wb.active
        filas_iter = list(ws.iter_rows(values_only=True))
        if not filas_iter:
            return [], []
        encabezados = [str(c) if c is not None else "" for c in filas_iter[0]]
        filas = [list(f) for f in filas_iter[1:]]
        return encabezados, filas

    @staticmethod
    def _leer_csv(path: Path) -> tuple[list[str], list[list]]:
        with open(path, newline="", encoding="utf-8-sig") as f:
            lector = list(csv.reader(f))
        if not lector:
            return [], []
        return lector[0], lector[1:]

    def guardar(self, path: str | Path) -> Path:
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        # normalizar el ancho de cada fila al de los encabezados
        ancho = len(self.encabezados)
        filas = [(f + [None] * ancho)[:ancho] for f in self.filas]
        if path.suffix.lower() in (".xlsx", ".xlsm"):
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Empresas"
            ws.append(self.encabezados)
            for f in filas:
                ws.append(f)
            wb.save(path)
        elif path.suffix.lower() == ".csv":
            with open(path, "w", newline="", encoding="utf-8-sig") as fh:
                w = csv.writer(fh)
                w.writerow(self.encabezados)
                w.writerows([["" if c is None else c for c in f] for f in filas])
        else:
            raise ValueError(f"Formato no soportado: {path.suffix} (usa .xlsx o .csv)")
        return path

    # ─── Mapeo de columnas ───────────────────────────────────────────────────

    def _mapear_columnas(self) -> dict[str, int]:
        norm = {_normalizar(h): i for i, h in enumerate(self.encabezados)}
        mapa: dict[str, int] = {}
        for campo, alias in ALIAS_COLUMNAS.items():
            for a in alias:
                idx = norm.get(_normalizar(a))
                if idx is not None:
                    mapa[campo] = idx
                    break
        return mapa

    def asegurar_columna(self, campo: str) -> int:
        """Devuelve el índice de la columna del campo; la crea si no existe."""
        if campo in self.mapa:
            return self.mapa[campo]
        encabezado = ENCABEZADOS_CANONICOS.get(campo) or ALIAS_COLUMNAS[campo][0].title()
        self.encabezados.append(encabezado)
        idx = len(self.encabezados) - 1
        for f in self.filas:
            while len(f) <= idx:
                f.append(None)
        self.mapa[campo] = idx
        return idx

    # ─── Acceso a celdas ─────────────────────────────────────────────────────

    def valor(self, fila: list, campo: str) -> object:
        idx = self.mapa.get(campo)
        if idx is None or idx >= len(fila):
            return None
        return fila[idx]

    def vacia(self, fila: list, campo: str) -> bool:
        return _vacio(self.valor(fila, campo))

    def set(self, fila: list, campo: str, valor: object, sobrescribir: bool = False) -> bool:
        """Escribe el valor sólo si la celda está vacía (o si ``sobrescribir``).

        Devuelve True si efectivamente escribió algo.
        """
        if _vacio(valor):
            return False
        idx = self.asegurar_columna(campo)
        while len(fila) <= idx:
            fila.append(None)
        if not sobrescribir and not _vacio(fila[idx]):
            return False
        fila[idx] = valor
        return True
